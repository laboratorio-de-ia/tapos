import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from models import Edital, Analise

# Modelo comercial fixo do produto: toda extração de edital gera exatamente
# esta planilha (uma única aba "Modelo"), idêntica em estrutura ao template —
# não um relatório genérico com abas extras. É o artefato que o time de
# precificação usa para orçar a licitação.
TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "Dom de Limpar - Modelo Precificacao.xlsx"

# Premissas de negócio do template (não vêm do edital, são a política de
# precificação padrão da empresa) — mantidas como constantes para ficar
# fácil de ajustar sem caçar números soltos no meio do código.
REGIME_TRIBUTARIO_PADRAO = "Lucro Presumido / operação padrão B2G"
ALIQUOTA_IMPOSTOS_PADRAO = 0.12
MARGEM_LIQUIDA_ALVO_PADRAO = 0.30
BASE_DE_COMPRA_PADRAO = "Grande BH"
ENTREGA_PADRAO = "Entrega parcial"

LINHA_PRIMEIRO_ITEM = 13


def _descricao_item(item) -> str:
    partes = [item.descricao or ""]
    if item.especificacoes:
        partes.append(item.especificacoes)
    texto = " - ".join(p for p in partes if p)
    return re.sub(r"\s+", " ", texto).strip()[:600]


def _copiar_estilo(origem, destino):
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.number_format = origem.number_format


def _ajustar_linhas_de_itens(ws, n_itens: int) -> int:
    """Redimensiona o bloco de itens do template (que vem pré-preenchido com
    os itens de um edital de exemplo) para caber exatamente `n_itens`,
    preservando a formatação (moeda, percentual) das colunas de precificação.
    Retorna a última linha de item usada (LINHA_PRIMEIRO_ITEM - 1 se vazio).
    """
    linha = LINHA_PRIMEIRO_ITEM
    while ws.cell(row=linha, column=1).value is not None:
        linha += 1
    n_itens_modelo = linha - LINHA_PRIMEIRO_ITEM

    if n_itens < n_itens_modelo:
        ws.delete_rows(LINHA_PRIMEIRO_ITEM + n_itens, n_itens_modelo - n_itens)
    elif n_itens > n_itens_modelo:
        linha_estilo = LINHA_PRIMEIRO_ITEM + max(n_itens_modelo - 1, 0)
        extras = n_itens - n_itens_modelo
        ws.insert_rows(LINHA_PRIMEIRO_ITEM + n_itens_modelo, extras)
        for offset in range(extras):
            linha_destino = LINHA_PRIMEIRO_ITEM + n_itens_modelo + offset
            for col in range(1, ws.max_column + 1):
                _copiar_estilo(ws.cell(row=linha_estilo, column=col), ws.cell(row=linha_destino, column=col))

    return LINHA_PRIMEIRO_ITEM + n_itens - 1 if n_itens else LINHA_PRIMEIRO_ITEM - 1


def _preencher_planilha_precificacao(wb, edital: Edital, analise: Analise) -> None:
    ws = wb["Modelo"]

    orgao = edital.orgao or "-"
    ws["A1"] = f"Análise Financeira de Edital - {orgao}"
    partes_subtitulo = [
        p
        for p in [
            edital.modalidade.title() if edital.modalidade else None,
            f"Edital {edital.numero}" if edital.numero else None,
            edital.objeto,
        ]
        if p
    ]
    partes_subtitulo.append(f"Base de compra: {BASE_DE_COMPRA_PADRAO}")
    ws["A2"] = " | ".join(partes_subtitulo)

    ws["B5"] = orgao
    ws["B6"] = edital.numero or "-"
    ws["B7"] = REGIME_TRIBUTARIO_PADRAO
    ws["B8"] = ALIQUOTA_IMPOSTOS_PADRAO
    ws["B9"] = edital.objeto or analise.resumo_executivo or "-"
    ws["B10"] = MARGEM_LIQUIDA_ALVO_PADRAO

    itens = analise.objetos_identificados or []
    ultima_linha_item = _ajustar_linhas_de_itens(ws, len(itens))

    for i, item in enumerate(itens):
        linha = LINHA_PRIMEIRO_ITEM + i
        ws.cell(row=linha, column=1, value=item.numero if item.numero is not None else i + 1)
        ws.cell(row=linha, column=2, value=_descricao_item(item) or "-")
        ws.cell(row=linha, column=3, value=item.quantidade)
        ws.cell(row=linha, column=4, value=(item.unidade or "-").upper())
        ws.cell(row=linha, column=5, value=item.modalidade_entrega or ENTREGA_PADRAO)
        # colunas F..P (embalagem/preços/impostos/margem) ficam em branco:
        # dependem de pesquisa de mercado que a extração do edital não
        # fornece — são preenchidas manualmente pelo time de precificação.
        for col in range(6, 17):
            ws.cell(row=linha, column=col, value=None)

    # localiza o bloco "Resumo Executivo de Precificação" logo após o bloco
    # de itens — a posição real depende de quantas linhas de item existem
    linha_resumo = None
    for row in ws.iter_rows(min_row=ultima_linha_item + 1, max_col=1):
        if row[0].value and "Resumo Executivo de Precificação" in str(row[0].value):
            linha_resumo = row[0].row
            break
    if linha_resumo is None:
        return

    faixa_ini = LINHA_PRIMEIRO_ITEM
    faixa_fim = ultima_linha_item if itens else LINHA_PRIMEIRO_ITEM
    linha_qtde, linha_compra, linha_venda, linha_lucro, linha_margem, linha_etp = (
        linha_resumo + i for i in range(1, 7)
    )

    ws.cell(row=linha_qtde, column=2, value=len(itens))
    if itens:
        ws.cell(row=linha_compra, column=2, value=f"=SUMPRODUCT(C{faixa_ini}:C{faixa_fim},H{faixa_ini}:H{faixa_fim})")
        ws.cell(row=linha_venda, column=2, value=f"=SUMPRODUCT(C{faixa_ini}:C{faixa_fim},M{faixa_ini}:M{faixa_fim})")
        ws.cell(row=linha_lucro, column=2, value=f"=SUMPRODUCT(C{faixa_ini}:C{faixa_fim},O{faixa_ini}:O{faixa_fim})")
    else:
        ws.cell(row=linha_compra, column=2, value=0)
        ws.cell(row=linha_venda, column=2, value=0)
        ws.cell(row=linha_lucro, column=2, value=0)
    ws.cell(row=linha_margem, column=2, value=f"=IFERROR(B{linha_lucro}/B{linha_venda},0)")

    # o edital raramente informa um valor total explícito (muitos usam
    # "orçamento sigiloso"); quando os itens trazem valor unitário estimado
    # pelo próprio edital, soma-se qtd × valor como aproximação do ETP
    valor_contratacao = sum(
        (item.valor_estimado or 0) * (item.quantidade or 0) for item in itens if item.valor_estimado
    )
    ws.cell(row=linha_etp, column=2, value=valor_contratacao if valor_contratacao else "-")


def gerar_excel(edital: Edital, analise: Analise, caminho_saida: str) -> str:
    """Gera a planilha de precificação a partir do template, com uma única
    aba ("Modelo") — estrutura idêntica ao arquivo
    `Dom de Limpar - Modelo Precificacao.xlsx` em toda extração."""
    wb = load_workbook(TEMPLATE_PATH)
    _preencher_planilha_precificacao(wb, edital, analise)

    for nome_aba in list(wb.sheetnames):
        if nome_aba != "Modelo":
            del wb[nome_aba]
    wb.active = 0

    Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return str(caminho_saida)
